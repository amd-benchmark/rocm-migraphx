from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import openpyxl
import pandas
import argparse
import pathlib
import os


#set execution arguments
parser = argparse.ArgumentParser()
parser.add_argument('-t', '--test', type=str, required=True, help="Path to test results")
parser.add_argument('-r', '--report', type=str, required=True, help="Path to test reports")
args = parser.parse_args()

#find path to lastest test result
find_last=max(pathlib.Path(args.test).glob('*/'), key=os.path.getmtime)
os.chdir(find_last)
print(f"Last test result: {find_last}")

#read lastest test result data
last_result = pandas.read_csv('results.csv', header=None)
with open('commit.txt') as git:
    commit = str(git.readlines(1))[9:15]
git.close()
with open('rocm.txt') as rocm:
    version = str(rocm.readlines()).split("/")[4].split("-")[1].split("\\")[0]
rocm.close()

#find lastest report and open
os.chdir(args.report)
last_report=max(pathlib.Path(args.report).glob('*/'), key=os.path.getmtime)
print(f"Last report: {last_report}")
wb = openpyxl.load_workbook(filename = last_report)
ws = wb.active

#copy data (this week to last week) 
for src, dst in zip(ws['C:C'], ws['E:E']):
    dst.value = src.value
print("Old data moved to last week")

#update cell values
ws.cell(row=1, column=3).value = 'ROCm ' + version
ws.cell(row=3, column=3).value = 'git: ' + commit
ws.cell(row=2, column=5).value = 'last week'

#import new test result
report = dataframe_to_rows(last_result, index=False, header=False)
for r_idx, row in enumerate(report, 5):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
print(f"New data imported from {find_last}")

#get current date
date = datetime.today().strftime('%Y-%m-%d')

#update sheet name
new_name = f'results-{date}'
ws.title = new_name 

#save as new report
new_report = f'{new_name}.xlsx'
wb.save(new_report)
print(f"New report saved: {new_report}")